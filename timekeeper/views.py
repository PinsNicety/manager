from django.shortcuts import redirect, render
from django.contrib.auth.models import User
from datetime import datetime, date
from . import models
from openpyxl import load_workbook
from django.http import HttpResponse

# Create your views here.
def home(request):
    """ Handles all the actions on the timekeeper main page. """

    content = {
        'left_header': 'home',
        'user': request.user
    }

    return render(request, 'timekeeper/timekeeper_sub.html', content)

def entries(request):
    """ Shows all entries for current user in the current pay period. """

    content = {
        'left_header': 'entries',
        'table_month': str(date.today().strftime("%B")),
    }

    if request.method == 'POST':
        time_in, time_out, total_time = set_times(request.POST['time_in'],
                                                request.POST['time_out'])


        entry = models.Entry.objects.create(site=request.POST['site'],
                                            time_in=time_in,
                                            time_out=time_out,
                                            comments=request.POST['comments'],
                                            total_time=total_time,
                                            user = request.user
        )

        content['entries'] = models.Entry.objects.filter(user=request.user, time_out__month=date.today().month).order_by('time_out')
        return render(request, 'timekeeper/entries.html', content)
    else:
        content['entries'] = models.Entry.objects.filter(user=request.user, time_out__month=date.today().month).order_by('time_out')
        return render(request, 'timekeeper/entries.html', content)

def entries_export(request, table_month):
    """Exports .xlxs files to user with entries for current month."""
    wb = load_workbook('media/walktest/detail_timesheet.xlsx')
    ws = wb.active
    ws['B3'] = f"WEEK ENDING: {table_month}"
    ws['B5'] = f"EMPLOYEE NAME: {request.user.username}"

    entries = models.Entry.objects.filter(user=request.user, time_out__month=date.today().month).order_by('time_out')

    row_data = set_entry_data(entries)
    row_counter = 11

    for row in row_data:
        ws[f"B{row_counter}"] = row['date']
        ws[f"C{row_counter}"] = row['day']
        ws[f"F{row_counter}"] = row['total']
        ws[f"H{row_counter}"] = row['customer']
        ws[f"K{row_counter}"] = row['remarks']
        row_counter += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=test_report.xlsx'
    wb.save(response)
    return response

def set_entry_data(entries):
    """
    Accepts objects from the view and sorts data to readable form for the
    detail timesheet.
    Data return as [{}]
    """
    row_data = []

    for entry in entries:
        if len(row_data) == 0:
            row_data.append({
                'date': entry.time_out.strftime("%-d-%b"),
                'day': entry.time_out.strftime("%a"),
                'total': entry.total_time,
                'customer': entry.site,
                'remarks': entry.comments
            })
        else:
            if entry.time_out.strftime("%-d-%b") == row_data[-1]['date']:
                row_data.append({
                    'date': '',
                    'day': '',
                    'total': entry.total_time,
                    'customer': entry.site,
                    'remarks': entry.comments
                })
            else:
                row_data.append({
                    'date': entry.time_out.strftime("%-d-%b"),
                    'day': entry.time_out.strftime("%a"),
                    'total': entry.total_time,
                    'customer': entry.site,
                    'remarks': entry.comments
                })

    return row_data

def expenses(request):
    """ Shows all expenses for current user in the current pay period. """

    content = {
        'left_header': 'expenses',
        'date_string': str(date.today()),
        'table_month': str(date.today().strftime("%B"))
    }

    if request.method == 'POST':

        new_date = set_date(request.POST['date'])
        expense = models.Expense.objects.create(site=request.POST['site'],
                                            miles=request.POST['miles'],
                                            date=new_date,
                                            user=request.user
        )
        content['expenses'] = models.Expense.objects.filter(user=request.user, date__month=date.today().month).order_by('date')
        return render(request, 'timekeeper/expenses.html', content)
    else:
        content['expenses'] = models.Expense.objects.filter(user=request.user, date__month=date.today().month).order_by('date')
        return render(request, 'timekeeper/expenses.html', content)

def expenses_export(request, table_month):
    """
    Exports .xlxs files to user with expenses for current month.
    """
    wb = load_workbook('media/walktest/expense_report.xlsx')
    ws = wb.active
    ws['C1'] = f"Name:____{request.user.username}____"
    ws['C2'] = f"Month:___{table_month}____"

    expenses = models.Expense.objects.filter(user=request.user, date__month=date.today().month).order_by('date')

    row_data = set_row_data(expenses)
    row_counter = 5

    for row in row_data:
        ws[f"B{row_counter}"] = row['date']
        ws[f"C{row_counter}"] = row['location']
        ws[f"F{row_counter}"] = row['miles']
        row_counter += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=test_report.xlsx'
    wb.save(response)
    return response

def set_row_data(expenses):
    """
    Accepts objects from view and sort data to readable form for the
    expense sheet.
    Data return as [{date, location, miles}]
    """
    row_data = []

    for expense in expenses:
        if len(row_data) == 0:
            row_data.append({
                'date': expense.date.strftime("%m/%d/%Y"),
                'location': expense.site,
                'miles': expense.miles
            })
        else:
            if expense.date.strftime("%m/%d/%Y") == row_data[-1]['date']:
                data = row_data.pop()
                data['location'] += f"/{expense.site}"
                data['miles'] += expense.miles
                row_data.append(data)
            else:
                row_data.append({
                    'date': expense.date.strftime("%m/%d/%Y"),
                    'location': expense.site,
                    'miles': expense.miles
                })
    return row_data

def set_times(time_in, time_out):
    """
    Takes in 2 date strings and returns them as datetime objects
    as well as the total time between the datetime objects.
    """

    time_in = datetime.strptime(time_in, "%Y-%m-%dT%H:%M")
    time_out = datetime.strptime(time_out, "%Y-%m-%dT%H:%M")

    diff = time_out - time_in
    total_time = diff.total_seconds() / 3600
    return time_in, time_out, total_time

def set_date(form_date):
    """ Takes in a date from the form and returns a DateTime object. """

    year, month, day = form_date.split("-")
    year = int(year)
    month = int(month)
    day = int(day)
    new_date = date(year, month, day)
    return new_date
